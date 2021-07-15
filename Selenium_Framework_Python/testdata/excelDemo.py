import openpyxl
book = openpyxl.load_workbook("C:\\Users\\Aparna\\PycharmProjects\\pytest\\pytestPackage\\Excel.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
print(sheet.cell(row=1, column=3).value)
print(sheet.cell(row=1, column=4).value)
sheet.cell(row=2, column=2).value = "Aparna"
print(sheet.cell(row=2, column=2).value)
print(sheet['C3'].value)
print(sheet.max_row)
print(sheet.max_column)
Dict = {}
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "TestCase2":
        for j in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print([Dict])





