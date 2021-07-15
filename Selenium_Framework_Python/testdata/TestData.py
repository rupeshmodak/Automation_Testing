import openpyxl


class TestData:
    testdatavariable = [{"name": "RM", "email": "rm@gmail.com", "gender": "Female"},
                        {"name": "AM", "email": "am@gmail.com", "gender": "Male"}]

    @staticmethod
    def get_TestData(get_test_data):
        book = openpyxl.load_workbook("C:\\Users\\Aparna\\PycharmProjects\\pytest\\pytestPackage\\Excel.xlsx")
        sheet = book.active
        dic = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == get_test_data:
                for j in range(2, sheet.max_column + 1):
                    dic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [dic]
